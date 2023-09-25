from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
max_row=ws.max_row
print(max_row)
for i in range(2,max_row+1):
    hours=float(ws['B'+str(i)].value)
    rate=ws['C'+str(i)].value
    if (type(hours)!=str and type(rate)!=str):
     salary=hours*rate    
     ws['D'+str(i)].value=salary
     print(salary)
     
wb.save('result.xlsx')
wb.close()
